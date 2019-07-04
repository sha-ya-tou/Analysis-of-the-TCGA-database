# -*- coding: utf-8 -*-
#将所有基因的度整理为一个表格
from openpyxl import load_workbook
import xlsxwriter

fname1=r"D:\Users\admin\Desktop\hub\list.xlsx"
fname2=r"D:\Users\admin\Desktop\hub\gene.xlsx"
#filename1=xlrd.open_workbook(fname1)
filename1=load_workbook(fname1)
#filename2=xlrd.open_workbook(fname2)
filename2=load_workbook(fname2)
#sheet1=filename1.sheet_by_index(0)
#sheet2=filename2.sheet_by_index(0)
sheetnames1 = filename1.get_sheet_names()#获取所有的sheet
sheet1=filename1.get_sheet_by_name(sheetnames1[0])#获取第一个sheet
sheetnames2 = filename2.get_sheet_names()
sheet2=filename2.get_sheet_by_name(sheetnames2[0])

filename=xlsxwriter.Workbook(r"D:\Users\admin\Desktop\hub\degree.xlsx")#新建一个文件
sheet=filename.add_worksheet('sheet1')#添加一个sheet

#nrows1=sheet1.nrows
nrows1=sheet1.max_row
#print "The first file line:%d" %nrows1
nrows2=sheet2.max_row
ncols2=sheet2.max_column
#print "The second file line:%d col:%d" %(nrows2,ncols2)

count=0
for i in range(1,nrows1+1):#第一行到最后一行
    left=sheet1.cell(i,1).value#第i行第一列的值
    right=sheet1.cell(i,2).value    
    for j in range(2,ncols2+1):
        if(left==sheet2.cell(j,1).value):
            for k in range(2,nrows2+1):
                if(right==sheet2.cell(1,k).value):
                    sheet.write(j-2,k-2,'1')
                    sheet.write(k-2,j-2,'1')#k-1行，j-1列
                    count+=2
                    continue
            continue
#print(sheet.cell(8,9).value)

#for i in range(1,nrows2+1):#第一行到最后一行   
    #for j in range(1,ncols2+1):
        #if(sheet.cell(i-1,j-1).value != 1):
            #sheet.write(i-1,j-1,'0')
print(count)
filename.close()
filename1.close()
filename2.close()

